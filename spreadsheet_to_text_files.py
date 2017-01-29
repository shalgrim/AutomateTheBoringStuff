"""
convert a spreadsheet to a directory of text files. Rows in Column A become
lines in file A.txt, rows in column B become lines in file B.txt, and so on
"""

import logging
import openpyxl
import os
from openpyxl.utils import get_column_letter
from srhpytools_srh.options.parsers import GenArgParser
from srhpytools_srh.util.mylogging import config_root_file_logger
from typing import List

logger = logging.getLogger('automate_boring.spreadsheet_to_text_files')


def strip_blank_lines(lines: List[str]) -> List[str]:
    """
    removes from lines consecutive elements at end that are only whitespace
    or empty
    :param lines: list of strings
    :return: copy of list where elements at end that are blank or empty are
    sliced off
    """
    for i in range(len(lines)-1, -1, -1):
        if lines[i] and lines[i].strip():
            return lines[:i+1]

    return []


def main(filename, outdir):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    os.makedirs(outdir, exist_ok=True)

    for col_idx in range(sheet.max_column):
        col = get_column_letter(col_idx+1)
        lines = []
        for row_idx in range(sheet.max_row):
            row = row_idx + 1
            lines.append(sheet[f'{col}{row}'].value)

        lines = strip_blank_lines(lines)
        lines = [f'{line}\n' if line is not None else '\n' for line in lines]
        with open(os.path.join(outdir, f'{col}.txt'), 'w') as f:
            f.writelines(lines)


if __name__ == '__main__':
    parser = GenArgParser(usage='%(prog)s -f filename [-d outdir] [options]')
    parser.add_argument('-f', '--filename', help='excel file to convert')
    parser.add_argument('-d', '--outdir', help='directory to write text '
                                               'files to')
    args = parser.parse_args()
    config_root_file_logger(args.logfile, args.loglevel, args.logmode)
    main(args.filename, args.outdir)
