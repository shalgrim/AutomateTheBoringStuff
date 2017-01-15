#! python 3
# update_produce.py - Corrects costs in produce sales spreadsheet

import openpyxl

wb = openpyxl.load_workbook('online_materials/produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# The produce types and their updated prices
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

# Loop through the rows and update the prices
for row_num in range(2, sheet.max_row): # skip the first row
    produce_name = sheet.cell(row=row_num, column=1)
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]

wb.save('online_materials/updatedProduceSales.xlsx')
