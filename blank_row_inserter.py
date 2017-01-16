"""
page 291
Takes two integers and a filename string as command line arguments. First
integer is N and second is M. Starting at row N, the program should insert M
blank rows into the spreadsheet.
"""

import logging
import openpyxl
from srhpytools_srh.options.parsers import GenArgParser
from srhpytools_srh.util.mylogging import config_root_file_logger

logger = logging.getLogger('automate_boring.blank_row_inserter')


def main(filename, start_row, rows_to_insert, outfile=None):
    if start_row < 1 or rows_to_insert < 0:
        return

    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    for row_idx in range(sheet.max_row, start_row - 1, -1):
        if row_idx % 1000 == 0:
            logger.info(f'copying line {row_idx}')
        for col_idx in range(1, sheet.max_column + 1):
            source_cell = sheet.cell(column=col_idx, row=row_idx)
            target_cell = sheet.cell(column=col_idx,
                                     row=row_idx + rows_to_insert)
            target_cell.value = source_cell.value

    for row_idx in range(start_row, start_row + rows_to_insert):
        for col_idx in range(1, sheet.max_column + 1):
            sheet.cell(column=col_idx, row=row_idx).value = None

    wb.save(outfile)


if __name__ == '__main__':
    parser = GenArgParser(usage='%(prog)s start_row num_new_rows filename')
    parser.add_argument('start_row', metavar='N', type=int,
                        help='the row at which to insert blank rows')
    parser.add_argument('num_new_rows', metavar='M', type=int,
                        help='the number of blank rows to insert')
    parser.add_argument('filename', help='the file to insert rows into')
    args = parser.parse_args()

    config_root_file_logger(args.logfile, args.loglevel, args.logmode)

    main(args.filename, args.start_row, args.num_new_rows, args.outfile)
