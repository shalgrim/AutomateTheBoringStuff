"""
page 291
Takes a number N from the command line and creates an NxN multiplication
table in an Excel spreadsheet
"""

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from srhpytools_srh.options.parsers import GenArgParser

if __name__ == '__main__':
    usage = '%(prog)s N [options]'
    parser = GenArgParser(usage=usage)
    parser.add_argument('table_size', metavar='N', type=int,
                        help='the size of one side of multiplication table')
    args = parser.parse_args()
    n = args.table_size

    wb = openpyxl.Workbook()
    sheet = wb.active
    bold_font = sheet['A1'].font.copy(bold=True)

    for i in range(1, n+1):
        sheet[f'A{i+1}'] = i
        sheet[f'A{i+1}'].font = bold_font
        sheet[f'{get_column_letter(i+1)}1'] = i
        sheet[f'{get_column_letter(i+1)}1'].font = bold_font

    for i in range(1, n+1):
        for j in range(1, n+1):
            sheet[f'{get_column_letter(i+1)}{j+1}'] = i*j

    wb.save(args.outfile)


