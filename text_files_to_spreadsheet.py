"""
reads in content of directory of text files, inserts them into spreadsheet
with one line of text per row. lines of first file go in column A, and so on
(page 293)
"""

import logging
import openpyxl
import os
from openpyxl.utils import get_column_letter
from srhpytools_srh.options.parsers import GenArgParser
from srhpytools_srh.util.mylogging import config_root_file_logger

logger = logging.getLogger('automate_boring.text_files_to_spreadsheet')


def main(dirname, outfile):
    file_lines = []
    for fn in os.listdir(dirname):
        with open(os.path.join(dirname, fn)) as f:
            file_lines.append([line.rstrip() for line in f.readlines()])

    wb = openpyxl.Workbook()
    sheet = wb.active
    for i, f in enumerate(file_lines):
        col = get_column_letter(i+1)
        for j, l in enumerate(f):
            row = j+1
            sheet[f'{col}{row}'] = l

    wb.save(outfile)


if __name__ == '__main__':
    parser = GenArgParser(usage='%(prog)s -d dirname [-o outfile] [options]')
    parser.add_argument('-d', '--dirname', help='directory of text files')
    args = parser.parse_args()
    config_root_file_logger(args.logfile, args.loglevel, args.logmode)
    main(args.dirname, args.outfile)



