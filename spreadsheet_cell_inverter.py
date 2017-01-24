"""
inverts rows and columns of cells of active sheet in workbook (page 292)
"""

import logging
import openpyxl
from openpyxl.utils import get_column_letter
from srhpytools_srh.options.parsers import GenArgParser
from srhpytools_srh.util.mylogging import config_root_file_logger

logger = logging.getLogger('automate_boring.blank_row_inserter')

def main(filename, outfile=None):

    # read data into matrix
    wb = openpyxl.load_workbook(filename)
    sheet: openpyxl.worksheet.worksheet.Worksheet = wb.active
    matrix = [[None]*sheet.max_column]*sheet.max_row
    inverted = [[None]*sheet.max_row]*sheet.max_column
    for column_idx in range(1, sheet.max_column+1):
        column = get_column_letter(column_idx)
        for row in range(1, sheet.max_row+1):
            matrix[row-1][column_idx-1] = sheet[f'{column}{row}'].value

    # invert matrix
    for row_idx in range(len(matrix)):
        for col_idx in range(len(matrix[row_idx])):
            inverted[col_idx][row_idx] = matrix[row_idx][col_idx]

    # clear out cells in sheet
    for column_idx in range(1, sheet.max_column+1):
        column = get_column_letter(column_idx)
        for row_idx in range(1, sheet.max_row+1):
            sheet[f'{column}{row_idx}'] = None

    # write data back to cells
    for row_idx in range(len(inverted)):
        for column_idx in range(len(inverted[row_idx])):
            column = get_column_letter(column_idx+1)
            sheet[f'{column}{row_idx+1}'] = inverted[row_idx][column_idx]

    if not outfile:
        logger.warning('will overwrite existing file')
    wb.save(outfile)


if __name__ == '__main__':
    parser = GenArgParser(usage='%(prog)s -f filename [-o outfile] [options]')
    parser.add_argument('-f', '--filename', help='file to invert')
    args = parser.parse_args()
    config_root_file_logger(args.logfile, args.loglevel, args.logmode)
    main(args.filename, args.outfile)


